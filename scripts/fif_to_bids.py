from __future__ import annotations

import argparse
import csv
import json
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import mne
from mne_bids import BIDSPath, make_dataset_description, write_raw_bids
from openpyxl import load_workbook


DEFAULT_SOURCE_DIR = Path("datas/data_01_ECoG_clean_1")
DEFAULT_BIDS_ROOT = Path("datas/data_02_BIDS")
DEFAULT_METADATA_XLSX = Path("datas/ele_pos.xlsx")
DEFAULT_DATASET_NAME = "Ecog Glioma Clean BIDS"
DEFAULT_TASK = "rest"
DEFAULT_SESSION_FILTER = "before"
DEFAULT_DATATYPE = "ieeg"
DEFAULT_CHANNEL_TYPE = "ecog"
DEFAULT_FORMAT = "EDF"
DEFAULT_LINE_FREQ = 50.0
DEFAULT_REFERENCE = "n/a"
DEFAULT_DESCRIPTION = "clean"

SUPPORTED_DATATYPES = ("auto", "eeg", "ieeg", "meg", "nirs")
SUPPORTED_CHANNEL_TYPES = ("preserve", "eeg", "ecog", "seeg", "dbs")
SUPPORTED_FORMATS = ("auto", "EDF", "BrainVision", "EEGLAB", "FIF")


@dataclass
class MetadataRow:
    patient_id: str
    subject: str
    session: str
    row_number: int


@dataclass
class ParsedName:
    stem: str
    patient_id: str
    run: str


@dataclass
class ConversionRecord:
    source_file: str
    status: str
    patient_id: str
    subject: str
    session: str
    run: str
    datatype: str
    output_file: str
    reason: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert cleaned FIF electrophysiology recordings into a BIDS "
            "derivative dataset."
        )
    )
    parser.add_argument(
        "--source-dir",
        type=Path,
        default=DEFAULT_SOURCE_DIR,
        help="Directory containing input .fif files.",
    )
    parser.add_argument(
        "--bids-root",
        type=Path,
        default=DEFAULT_BIDS_ROOT,
        help="Root directory of the output BIDS dataset.",
    )
    parser.add_argument(
        "--metadata-xlsx",
        type=Path,
        default=DEFAULT_METADATA_XLSX,
        help="Spreadsheet containing patient_id / Sub-ID / sesion mapping.",
    )
    parser.add_argument(
        "--session-filter",
        default=DEFAULT_SESSION_FILTER,
        help="Only export rows matching this sesion value from the spreadsheet.",
    )
    parser.add_argument(
        "--datatype",
        choices=SUPPORTED_DATATYPES,
        default=DEFAULT_DATATYPE,
        help="Target BIDS datatype. Use auto to infer from source data.",
    )
    parser.add_argument(
        "--channel-type",
        choices=SUPPORTED_CHANNEL_TYPES,
        default=DEFAULT_CHANNEL_TYPE,
        help="Override all source data channels to this type, or preserve them.",
    )
    parser.add_argument(
        "--task",
        default=DEFAULT_TASK,
        help="BIDS task label to assign to exported recordings.",
    )
    parser.add_argument(
        "--format",
        dest="export_format",
        choices=SUPPORTED_FORMATS,
        default=DEFAULT_FORMAT,
        help="Target file format for BIDS payload files.",
    )
    parser.add_argument(
        "--line-freq",
        type=float,
        default=DEFAULT_LINE_FREQ,
        help="Power line frequency to set when missing.",
    )
    parser.add_argument(
        "--reference",
        default=DEFAULT_REFERENCE,
        help="Reference string to write into iEEG/EEG sidecars when applicable.",
    )
    parser.add_argument(
        "--description",
        default=DEFAULT_DESCRIPTION,
        help="BIDS desc entity to mark preprocessed recordings.",
    )
    parser.add_argument(
        "--dataset-name",
        default=DEFAULT_DATASET_NAME,
        help="Name to write into dataset_description.json.",
    )
    parser.add_argument(
        "--dataset-license",
        default="n/a",
        help="License to write into dataset_description.json.",
    )
    parser.add_argument(
        "--generated-by-name",
        default="fif_to_bids.py",
        help="GeneratedBy.Name value for dataset_description.json.",
    )
    parser.add_argument(
        "--source-datasets-url",
        default="n/a",
        help="SourceDatasets[0].URL value for dataset_description.json.",
    )
    parser.add_argument(
        "--validator-cmd",
        nargs="+",
        help=(
            "Explicit validator command. Example: --validator-cmd "
            ".venv\\Scripts\\bids-validator-deno.exe"
        ),
    )
    parser.add_argument(
        "--validate",
        action="store_true",
        help="Run bids-validator after export if available.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing BIDS files if they already exist.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse metadata and report planned actions without writing files.",
    )
    return parser.parse_args()


def repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def as_repo_path(path: Path) -> Path:
    return path if path.is_absolute() else repo_root() / path


def normalize_subject(value: str) -> str:
    subject = value.strip()
    if subject.lower().startswith("sub-"):
        subject = subject[4:]
    return subject


def load_session_metadata(metadata_xlsx: Path, session_filter: str) -> dict[str, MetadataRow]:
    workbook = load_workbook(metadata_xlsx, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    indexes = {name: index for index, name in enumerate(header) if name is not None}
    required_columns = ("patient_id", "Sub-ID", "sesion")
    missing_columns = [name for name in required_columns if name not in indexes]
    if missing_columns:
        missing_display = ", ".join(missing_columns)
        raise ValueError(f"Missing required metadata columns: {missing_display}")

    mapping: dict[str, MetadataRow] = {}
    for row_number, row in enumerate(rows, start=2):
        patient_value = row[indexes["patient_id"]]
        subject_value = row[indexes["Sub-ID"]]
        session_value = row[indexes["sesion"]]
        if patient_value is None or subject_value is None or session_value is None:
            continue

        session = str(session_value).strip()
        if session != session_filter:
            continue

        patient_id = str(patient_value).strip()
        mapping[patient_id] = MetadataRow(
            patient_id=patient_id,
            subject=normalize_subject(str(subject_value)),
            session=session,
            row_number=row_number,
        )
    return mapping


def parse_source_name(path: Path) -> ParsedName:
    parts = path.stem.split("_")
    if len(parts) < 3:
        raise ValueError(
            "Expected filename pattern YYYYMMDD_patientid_run.fif, "
            f"got {path.name!r}"
        )

    patient_id = parts[1].strip()
    run_token = parts[2].strip()
    if not patient_id or not run_token:
        raise ValueError(f"Could not parse patient_id/run from {path.name!r}")

    run = "".join(character for character in run_token if character.isdigit()) or run_token
    return ParsedName(stem=path.stem, patient_id=patient_id, run=run)


def resolve_validator_command(explicit_command: list[str] | None) -> list[str] | None:
    if explicit_command:
        return explicit_command

    candidates = [
        [str(repo_root() / ".venv" / "Scripts" / "bids-validator-deno.exe")],
        ["bids-validator-deno"],
        ["bids-validator"],
    ]
    for command in candidates:
        executable = command[0]
        if Path(executable).exists() or shutil.which(executable):
            return command
    return None


def expected_payload_extension(export_format: str) -> str | None:
    extension_map = {
        "EDF": ".edf",
        "BrainVision": ".vhdr",
        "EEGLAB": ".set",
        "FIF": ".fif",
    }
    return extension_map.get(export_format)


def infer_datatype(raw: mne.io.BaseRaw) -> str:
    channel_types = set(raw.get_channel_types(unique=True))
    if "ecog" in channel_types or "seeg" in channel_types or "dbs" in channel_types:
        return "ieeg"
    if "eeg" in channel_types:
        return "eeg"
    if "fnirs_cw_amplitude" in channel_types or "hbo" in channel_types or "hbr" in channel_types:
        return "nirs"
    if "mag" in channel_types or "grad" in channel_types:
        return "meg"
    raise ValueError(f"Could not infer BIDS datatype from channel types: {sorted(channel_types)}")


def set_channel_type_override(raw: mne.io.BaseRaw, channel_type: str) -> None:
    if channel_type == "preserve":
        return
    mapping = {name: channel_type for name in raw.ch_names}
    raw.set_channel_types(mapping, on_unit_change="ignore")


def make_bids_path(
    bids_root: Path,
    metadata: MetadataRow,
    parsed: ParsedName,
    task: str,
    description: str,
    datatype: str,
) -> BIDSPath:
    return BIDSPath(
        subject=metadata.subject,
        session=metadata.session,
        task=task,
        run=parsed.run.zfill(2),
        description=description,
        root=bids_root,
        datatype=datatype,
    )


def ensure_dataset_description(
    bids_root: Path,
    dataset_name: str,
    dataset_license: str,
    generated_by_name: str,
    source_datasets_url: str,
    overwrite: bool,
) -> None:
    make_dataset_description(
        path=bids_root,
        name=dataset_name,
        dataset_type="derivative",
        data_license=dataset_license,
        generated_by=[{"Name": generated_by_name}],
        source_datasets=[{"URL": source_datasets_url}],
        overwrite=overwrite,
    )


def update_json(path: Path, updates: dict[str, Any]) -> None:
    data = json.loads(path.read_text(encoding="utf-8"))
    data.update(updates)
    path.write_text(json.dumps(data, indent=4) + "\n", encoding="utf-8")


def update_coordsystem_json(path: Path) -> None:
    updates = {
        "iEEGCoordinateProcessingDescription": (
            "Electrode coordinates have not been derived yet; "
            "placeholders remain until manual localization is completed."
        ),
        "iEEGCoordinateProcessingReference": "n/a",
    }
    update_json(path, updates)


def update_recording_sidecar(path: Path, datatype: str, reference: str) -> None:
    updates: dict[str, Any] = {}
    if datatype == "ieeg":
        updates["iEEGReference"] = reference
    elif datatype == "eeg":
        updates["EEGReference"] = reference
    if updates:
        update_json(path, updates)


def resolve_payload_path(written_bids_path: BIDSPath, datatype: str, export_format: str) -> Path:
    explicit_extension = expected_payload_extension(export_format)
    if explicit_extension is not None:
        return written_bids_path.copy().update(
            suffix=datatype,
            extension=explicit_extension,
        ).fpath

    candidates = sorted(
        written_bids_path.directory.glob(f"{written_bids_path.basename}_{datatype}.*")
    )
    if candidates:
        return candidates[0]

    return written_bids_path.copy().update(suffix=datatype).fpath


def write_csv_report(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_json_report(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=4) + "\n", encoding="utf-8")


def ensure_bidsignore(bids_root: Path) -> None:
    bidsignore_path = bids_root / ".bidsignore"
    entries = {"reports", "reports/", "reports/**"}
    if bidsignore_path.exists():
        existing = {
            line.strip()
            for line in bidsignore_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        }
        entries.update(existing)
    bidsignore_path.write_text(
        "\n".join(sorted(entries)) + "\n",
        encoding="utf-8",
    )


def run_validator(validator_command: list[str], bids_root: Path) -> dict[str, Any]:
    command = [*validator_command, str(bids_root)]
    completed = subprocess.run(
        command,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
    )
    stdout = completed.stdout
    stderr = completed.stderr
    has_error = "[ERROR]" in stdout or "[ERROR]" in stderr
    has_warning = "[WARNING]" in stdout or "[WARNING]" in stderr
    if has_error:
        status = "failed"
    elif has_warning:
        status = "passed_with_warnings"
    else:
        status = "passed"
    return {
        "status": status,
        "command": command,
        "returncode": completed.returncode,
        "stdout": stdout,
        "stderr": stderr,
    }


def print_summary(records: list[ConversionRecord], validation_report: dict[str, Any] | None) -> None:
    converted = [row for row in records if row.status == "converted"]
    planned = [row for row in records if row.status == "planned"]
    skipped = [row for row in records if row.status == "skipped"]
    print(f"Converted: {len(converted)}")
    if planned:
        print(f"Planned: {len(planned)}")
    print(f"Skipped: {len(skipped)}")
    if skipped:
        print("Skipped files:")
        for row in skipped:
            print(f"  - {row.source_file}: {row.reason}")
    if validation_report:
        print(f"Validation: {validation_report['status']}")
        print(f"Validator return code: {validation_report.get('returncode', 'n/a')}")


def main() -> int:
    args = parse_args()

    source_dir = as_repo_path(args.source_dir)
    bids_root = as_repo_path(args.bids_root)
    metadata_xlsx = as_repo_path(args.metadata_xlsx)
    reports_dir = bids_root / "reports"

    if not source_dir.exists():
        print(f"Source directory does not exist: {source_dir}", file=sys.stderr)
        return 1
    if not metadata_xlsx.exists():
        print(f"Metadata spreadsheet does not exist: {metadata_xlsx}", file=sys.stderr)
        return 1

    mapping = load_session_metadata(metadata_xlsx, args.session_filter)
    source_files = sorted(source_dir.glob("*.fif"))
    if not source_files:
        print(f"No .fif files found under {source_dir}", file=sys.stderr)
        return 1

    validator_command = resolve_validator_command(args.validator_cmd)
    records: list[ConversionRecord] = []

    if not args.dry_run:
        bids_root.mkdir(parents=True, exist_ok=True)
        ensure_dataset_description(
            bids_root=bids_root,
            dataset_name=args.dataset_name,
            dataset_license=args.dataset_license,
            generated_by_name=args.generated_by_name,
            source_datasets_url=args.source_datasets_url,
            overwrite=args.overwrite or not (bids_root / "dataset_description.json").exists(),
        )
        ensure_bidsignore(bids_root)

    for source_file in source_files:
        try:
            parsed = parse_source_name(source_file)
        except ValueError as error:
            records.append(
                ConversionRecord(
                    source_file=source_file.name,
                    status="skipped",
                    patient_id="",
                    subject="",
                    session="",
                    run="",
                    datatype="",
                    output_file="",
                    reason=str(error),
                )
            )
            continue

        metadata = mapping.get(parsed.patient_id)
        if metadata is None:
            records.append(
                ConversionRecord(
                    source_file=source_file.name,
                    status="skipped",
                    patient_id=parsed.patient_id,
                    subject="",
                    session=args.session_filter,
                    run=parsed.run.zfill(2),
                    datatype="",
                    output_file="",
                    reason=(
                        f"No metadata row with patient_id={parsed.patient_id} "
                        f"and sesion={args.session_filter}"
                    ),
                )
            )
            continue

        raw = mne.io.read_raw_fif(source_file, preload=False, verbose="ERROR")
        set_channel_type_override(raw, args.channel_type)
        datatype = infer_datatype(raw) if args.datatype == "auto" else args.datatype
        if raw.info.get("line_freq") in (None, 0):
            raw.info["line_freq"] = args.line_freq
        bids_path = make_bids_path(
            bids_root=bids_root,
            metadata=metadata,
            parsed=parsed,
            task=args.task,
            description=args.description,
            datatype=datatype,
        )

        output_file = str(
            resolve_payload_path(
                written_bids_path=bids_path,
                datatype=datatype,
                export_format=args.export_format,
            )
        )
        if args.dry_run:
            records.append(
                ConversionRecord(
                    source_file=source_file.name,
                    status="planned",
                    patient_id=parsed.patient_id,
                    subject=metadata.subject,
                    session=metadata.session,
                    run=parsed.run.zfill(2),
                    datatype=datatype,
                    output_file=output_file,
                    reason="",
                )
            )
            continue

        written_bids_path = write_raw_bids(
            raw=raw,
            bids_path=bids_path,
            format=args.export_format,
            overwrite=args.overwrite,
            verbose="ERROR",
        )

        sidecar_json = written_bids_path.copy().update(
            suffix=datatype,
            extension=".json",
        ).fpath
        update_recording_sidecar(sidecar_json, datatype=datatype, reference=args.reference)

        if datatype == "ieeg":
            coordsystem_json = written_bids_path.copy().update(
                task=None,
                run=None,
                description=None,
                suffix="coordsystem",
                extension=".json",
            ).fpath
            update_coordsystem_json(coordsystem_json)

        payload_path = resolve_payload_path(
            written_bids_path=written_bids_path,
            datatype=datatype,
            export_format=args.export_format,
        )
        records.append(
            ConversionRecord(
                source_file=source_file.name,
                status="converted",
                patient_id=parsed.patient_id,
                subject=metadata.subject,
                session=metadata.session,
                run=parsed.run.zfill(2),
                datatype=datatype,
                output_file=str(payload_path),
                reason="",
            )
        )

    report_rows = [
        {
            "source_file": row.source_file,
            "status": row.status,
            "patient_id": row.patient_id,
            "subject": row.subject,
            "session": row.session,
            "run": row.run,
            "datatype": row.datatype,
            "output_file": row.output_file,
            "reason": row.reason,
        }
        for row in records
    ]
    fieldnames = [
        "source_file",
        "status",
        "patient_id",
        "subject",
        "session",
        "run",
        "datatype",
        "output_file",
        "reason",
    ]
    conversion_report_path = reports_dir / "conversion_report.csv"
    validation_report: dict[str, Any] | None = None
    validation_report_path = reports_dir / "validation_report.json"

    if not args.dry_run:
        write_csv_report(conversion_report_path, report_rows, fieldnames)

        if args.validate:
            if validator_command is None:
                validation_report = {
                    "status": "unavailable",
                    "command": [],
                    "reason": "No bids-validator command could be resolved.",
                }
            else:
                validation_report = run_validator(validator_command, bids_root)
            write_json_report(validation_report_path, validation_report)

    print_summary(records, validation_report)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
